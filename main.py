from bleach import clean
import xlrd
import ipdb
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import args

class LatexRender:
    def __init__(self,configure:args.Configure) -> None:
        with open(configure.tableTemplate()) as f:
            self.__tableTemplate = f.read()
    def renderTable(self,colNumber:int,content:str,caption:str,label:str)->str:
        col = "{"
        for i in range(colNumber-1):
            col += "l|"
        col+="l}"
        return self.__format(self.__tableTemplate,col,content,
            "{"+caption+"}","{"+label+"}")
    def __format(self,template:str,*args)->str:
        return template.format(args)

class Transformer:
    def __init__(self,render:LatexRender) -> None:
        self.render = render

    def transWorkbook(self,configure:args.Configure)->str:
        return ""

class XlsTransformer(Transformer):
    def __init__(self,render:LatexRender) -> None:
        super().__init__(render)
        pass
    def __transCell(self,cell)->str:
        return str(cell.value)

    def __transRow(self,row)->str:
        res = self.__transCell(row[0])
        for cell in row[1:]:
            res += f" & {self.__transCell(cell)}"
        res += "\\\\"
        return res

    def __transSheet(self,sheet)->str:
        ipdb.set_trace()
        colNumber = sheet.ncols
        rowNumber = sheet.nrows
        content = []
        for index in range(rowNumber):
            content.append(self.__transRow(sheet.row(index)))
        return self.render.renderTable(colNumber,"\n".join(content),"","")

    def transWorkbook(self,configure:args.Configure) -> str:
        print("暂不支持")
        return ""
        source = xlrd.open_workbook(configure.sourceFile())
        latexSheets = []
        if configure.sheets() == None:
            for index,sheet in enumerate(source.sheets()):
                latexSheets.append(self.__transSheet(sheet))
        else:
            for index in configure.sheets():
                latexSheets.append(self.__transSheet(source.sheet_by_index(index)))
        return "\n".join(latexSheets)

class XlsxTransformer(Transformer):
    def __init__(self,render:LatexRender)->None:
        super().__init__(render)

    def __transCell(self,cell)->str:
        if cell.value is None:
            return ""
        return str(cell.value)

    def __transRow(self,row)->str:
        res = self.__transCell(row[0])
        for cell in row[1:]:
            res += f" & {self.__transCell(cell)}"
        res += "\\\\"
        return res

    def __transSheet(self,sheet:Worksheet)->str:
        mergedCells:list = sheet.merged_cells.ranges
        mergedCells.sort(key=lambda cell:cell.min_row)
        colNumber = sheet.max_column
        content = []
        mergedIndex = 0
        for index,row in enumerate(sheet.rows):
            if mergedCells[mergedIndex].min_row > index+1:
                content.append(self.__transRow(row) + "\\hline")
            elif mergedCells[mergedIndex].max_row == index+1:
                content.append(self.__transRow(row) + "\\hline")
                mergedIndex += 1
            else:
                content.append(self.__transRow(row))
        return self.render.renderTable(colNumber,"\n".join(content),"","")

    def transWorkbook(self,configure:args.Configure)->str:
        wb = openpyxl.load_workbook(configure.sourceFile())
        latexSheets = []
        if configure.sheets() == None:
            for sheet in wb:
                latexSheets.append(self.__transSheet(sheet))
        else:
            listIndex = 0
            for sheetIndex,sheet in enumerate(wb):
                if sheetIndex != configure.sheets()[listIndex]:
                    continue
                listIndex += 1
                latexSheets.append(self.__transSheet(sheet))
        wb.close()
        return "\n".join(latexSheets)

class ExcelTransformer:
    def __init__(self,configure:args.Configure,render:LatexRender) -> None:
        self.configure = configure
        if configure.fileType() == args.ExcelFileType.Xls:
            self.__transformer = XlsTransformer(render)
        elif configure.fileType() == args.ExcelFileType.Xlsx:
            self.__transformer = XlsxTransformer(render)
    def transWorkbook(self):
        if configure.outFile() == None:
            print(self.__transformer.transWorkbook(self.configure))
            return
        with open(configure.outFile(),'w',encoding='utf-8') as f:
            f.write(self.__transformer.transWorkbook(self.configure))
        
configure = args.Configure()
render = LatexRender(configure)
excelTransformer = ExcelTransformer(configure,render)
excelTransformer.transWorkbook()
